import openpyxl
import openpyxl as xl 
from openpyxl import Workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment
from openpyxl.cell import Cell
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors, Fill, fills
import random
import os
#Load  workbook
workbook = openpyxl.load_workbook('out.xlsx')
worksheet_out = workbook.worksheets[0]
worksheet_out = workbook.active
worksheet = workbook.worksheets[1]
mr = worksheet.max_row
mc = worksheet.max_column
#get time
#1 Get num_les
col_stt = 'A2'
i = 0
while (worksheet[col_stt].value is not None ):
    i= i+1
    col_stt = 'A' +  str(i+1)
    cell_null = worksheet[col_stt].value
num_les=i+1
#2 pretreatment

#3 get timeline
for i in range (2,num_les) :
    coler0 = 'J'+ str(i) # Subject

    colerGetLine_Day= 'H'+ str(i)
    colerGetLine_Time= 'I'+ str(i)
    colerGetLine_Add= 'J'+ str(i)

    colerLine1_Day= 'K'+ str(i) # Get Day
    colerLine1_Time= 'L'+ str(i) #Get time
    colerLine1_Add= 'M'+ str(i) #Get address

    colerLine2_Day= 'N'+ str(i) #Get day
    colerLine2_Time= 'O'+ str(i) #Get time
    colerLine2_Add= 'P'+ str(i) #Get address
    
    worksheet[colerLine1_Day] = '=LEFT('+colerGetLine_Day+',SEARCH(",",'+colerGetLine_Day+')-1)'
    worksheet[colerLine1_Time] = '=LEFT('+colerGetLine_Time+',SEARCH(",",'+colerGetLine_Time+')-1)'
    worksheet[colerLine1_Add] = '=LEFT('+colerGetLine_Add+', SEARCH("T",'+colerGetLine_Add+')+1)'

    worksheet[colerLine2_Day] = '=RIGHT('+colerGetLine_Day+',LEN('+colerGetLine_Day+')-SEARCH(",",'+colerGetLine_Day+'))'
    worksheet[colerLine2_Time] = '=RIGHT('+colerGetLine_Time+',LEN('+colerGetLine_Time+')-SEARCH(",",'+colerGetLine_Time+'))'
    worksheet[colerLine2_Add] = '=RIGHT('+colerGetLine_Add+', SEARCH("T",'+colerGetLine_Add+')+1)'
workbook.save("out.xlsx")
os.system('python convertFormulatoText.py')
#ReLoad  workbook
workbook = openpyxl.load_workbook('out.xlsx')
worksheet = workbook.worksheets[1]
for i in range (2,num_les) :
    coler1 = 'Q'+ str(i) #Line1
    coler2 = 'R'+ str(i) #Line1
    coler3 = 'S'+ str(i) #Line1

    coler4 = 'T'+ str(i) #Line2 
    coler5 = 'U'+ str(i) #Line2
    coler6 = 'V'+ str(i) #Line2

    colerLine1_Time= 'L'+ str(i) #Get time
    colerLine2_Time= 'O'+ str(i) #Get time2

    if worksheet.cell(row = i, column = 1 ).value != 0 :
        worksheet[coler1] = '=MONTH(' + colerLine1_Time + ')'
        worksheet[coler2] = '=DAY(' + colerLine1_Time + ')'
        worksheet[coler3] = '=C'+str(i)+'&CHAR(10)&M'+str(i)

        worksheet[coler4] = '=MONTH(' + colerLine2_Time + ')'
        worksheet[coler5] = '=DAY(' + colerLine2_Time + ')'
        worksheet[coler6] = '=C'+str(i)+'&CHAR(10)&P'+str(i)
    else:
        coler0 = 'I'+ str(i)
        coler1 = 'K'+ str(i)
        coler2 = 'L'+ str(i) 
        coler3 = 'M'+ str(i) 
        worksheet[coler1] = '=MONTH(' + coler0 + ')'
        worksheet[coler2] = '=DAY(' + coler0 + ')'
        worksheet[coler3] = '=C'+str(i)+'&CHAR(10)&J'+str(i)
workbook.save("out.xlsx")
os.system('python convertFormulatoText.py')
#ReLoad  workbook
workbook = openpyxl.load_workbook('out.xlsx')
worksheet = workbook.worksheets[1]
#Merge table
mr0 = worksheet_out.max_row
mc0 = worksheet_out.max_column
## identify for col 
for i in range (4, 10):
    workbook.worksheets[0].cell(row = 111, column = i ).value = 'T' + str(i-2)
workbook.worksheets[0].cell(row = 111, column = 10 ).value = 'CN'
## Merging
data_color = ['E0699C','E0D675','C25DE0','48E04E','7253E0','69ADE0','E07875','5DE0D3','E0A448','53E082']
for row in  range (2, num_les):
    if worksheet.cell(row = row, column = 1 ).value != 0 :
        scell = worksheet.cell(row = row, column = 17 ).value +3 #17
        ecell =  worksheet.cell(row = row, column = 18 ).value +3 #18
        col_day = worksheet.cell(row = row, column = 11 ).value #11

        if workbook.worksheets[0].cell(row = 9, column = 2).value == 'RELAX':
            if scell >= 9:
                ecell +=1
                scell +=1
        for checker in range (4,10):
            cell_checker = worksheet_out.cell(row = 111, column = checker).value
            if  cell_checker == col_day :
                workbook.worksheets[0].merge_cells(start_row= scell, start_column= checker, end_row= ecell , end_column= checker)
                workbook.worksheets[0].cell(row = scell, column = checker).value = workbook.worksheets[1].cell(row = row, column = 19).value #19
                top_left_cell = workbook.worksheets[0].cell(row = scell, column = checker)
                top_left_cell.fill = PatternFill("gray125", fgColor= random.choice(data_color))
#solve col2
for row in  range (2, num_les):
    if worksheet.cell(row = row, column = 1 ).value != 0 :
        scell = worksheet.cell(row = row, column = 20 ).value +3
        ecell =  worksheet.cell(row = row, column = 21 ).value +3
        col_day = worksheet.cell(row = row, column = 14 ).value

        if workbook.worksheets[0].cell(row = 9, column = 2).value == 'RELAX':
            if scell >= 9:
                ecell +=1
                scell +=1
        for checker in range (4,10):
            cell_checker = worksheet_out.cell(row = 111, column = checker).value
            if  cell_checker == col_day :
                workbook.worksheets[0].merge_cells(start_row= scell, start_column= checker, end_row= ecell , end_column= checker)
                workbook.worksheets[0].cell(row = scell, column = checker).value = workbook.worksheets[1].cell(row = row, column = 22).value
                top_left_cell = workbook.worksheets[0].cell(row = scell, column = checker)
                top_left_cell.fill = PatternFill("gray125", fgColor= random.choice(data_color))
#solve other:
for row in  range (2, num_les):
    if worksheet.cell(row = row, column = 1 ).value == 0 :
        try:
            scell = worksheet.cell(row = row, column = 11 ).value +3
            ecell =  worksheet.cell(row = row, column = 12 ).value +3
            col_day = worksheet.cell(row = row, column = 8 ).value

            if workbook.worksheets[0].cell(row = 9, column = 2).value == 'RELAX':
                if scell >= 9:
                    ecell +=1
                    scell +=1
            for checker in range (4,10):
                cell_checker = worksheet_out.cell(row = 111, column = checker).value
                if  cell_checker == col_day :
                    workbook.worksheets[0].merge_cells(start_row= scell, start_column= checker, end_row= ecell , end_column= checker)
                    workbook.worksheets[0].cell(row = scell, column = checker).value = workbook.worksheets[1].cell(row = row, column = 13).value
                    top_left_cell = workbook.worksheets[0].cell(row = scell, column = checker)
                    top_left_cell.fill = PatternFill("gray125", fgColor= random.choice(data_color))
        

        except:
            print("good for health")

for row in  range (2, num_les):
    if(workbook.worksheets[1].cell(row = row, column = 11).value is not None):
        workbook.worksheets[1].cell(row = row, column = 1).value =0
os.system('python convertFormulatoText.py')
#save
workbook.save("out.xlsx")
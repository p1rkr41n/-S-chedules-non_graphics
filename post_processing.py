# processing column non-lesson
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# create Workbook object

wb = openpyxl.load_workbook("./out.xlsx")
# set file path
ws = wb.worksheets[0]
ws= wb.active
ws2= wb.worksheets[1]
#process columns

for col in range(10,4,-1):
    check = 0
    for row in range(4,17):
        #print(ws.cell(row = row, column = col).value)
        if(ws.cell(row = row, column = col).value is not None):
            check +=1
            find = True 
            break 
    try:            # <-- break here too
        if find: 
            break
    except: 
        if check == 0:
            ws.column_dimensions[get_column_letter(col)].hidden= True              #hidden col éo dùng

#process rows non-lesson
last_less = 1
for i in range(2,10):
    try:
        if (ws2.cell(row = i, column = 12).value >= last_less) :
            last_less = ws2.cell(row = i, column = 12).value
    except:
        print("Error delete row!")
for i in range(2,10):
    try:
        if (ws2.cell(row = i, column = 21).value >= last_less) :
            last_less = ws2.cell(row = i, column = 21).value
    except:
        print("Error delete row!")
for i in range(2,10):
    try:
        if (ws2.cell(row = i, column = 18).value >= last_less) :
            last_less = ws2.cell(row = i, column = 18).value
    except:
        print("Error delete row!")
if wb.worksheets[0].cell(row = 9, column = 2).value == 'RELAX':
    last_less = last_less + 5
else:
    last_less = last_less + 4
for row in range(last_less,18):
    ws.row_dimensions[row].hidden= True              #hidden row éo dùng

# save workbook 
wb.save("./out.xlsx")
wb.close()